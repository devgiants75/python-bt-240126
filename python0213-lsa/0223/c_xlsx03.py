### pandas & openpyxl ###

# 파이썬에 생성한 데이터를 활용하여
# , 엑셀 파일 생성 - pandas
import pandas as pd

data = {
    'Name': ['a', 'b', 'c'],
    'Age': [28, 30, 29],
    'City': ['부산', '대전', '서울']
}

# 데이터프레임 생성
# : pandas의 DataFrame() 함수 사용
df = pd.DataFrame(data)

# 데이터 프레임을 엑셀 파일로 저장
# : 데이터프레임변수명.to_excel('엑셀파일명.xlsx', index=False)
# index=False
# : pandas가 엑셀 파일로 데이터를 저장할 때, 인덱스를 파일에 포함X
df.to_excel('example.xlsx', index=False)

# pandas 라이브러리 실행 시 사용하고 있는 엑셀 파일이 열려 있는 경우
# : permission Error가 발생

### openpyxl - 파이썬에서 엑셀 데이터 다루기 ###
# openpyxl 라이브러리의 load_workbook 모듈 사용
# : 지정된 경로에 있는 엑셀 파일을 불러와서
# : 파이썬에서 다룰 수 있는 객체로 변경
from openpyxl import load_workbook

# 엑셀 파일 불러오기
# load_workbook('파일명.확장자')
wb = load_workbook('example.xlsx')
print(wb) # 엑셀 파일이 담긴 메모리 주소값이 출력

# 기존에 있는 첫 번째 시트 선택
sheet = wb.active
print(sheet)

# 특정 이름의 시트 선택
# selected_sheet = wb['시트명'] # wb['Sheet1']

# 특정 위치의 시트에 접근
# 위치: 전체 시트가 리스트로 저장 - 인덱스 번호
# selected_sheet = wb.worksheets['위치번호(숫자)']
# wb.worksheets[1] - sheet2

# openpyxl을 사용하여 엑셀의 셀 서식, 폰트 및 색상 변경(디자인)
# : openpyxl의 스타일 기능을 사용
from openpyxl.styles import Font, Color, Alignment, PatternFill
# Font-글꼴 스타일
# Color-색상
# Alignment-배치(정렬)
# PatternFill-셀 배경 채우기

# 첫 번째 행(제목 행)의 폰트와 정렬 변경
# : 첫 번째 행에 있는 모든 셀에 대한 반복
for cell in sheet['1:1']:
    # 각 셀의 스타일 지정
    # cell변수명.font = Font모듈(속성 지정)
    # bold-두께, color-색상(16진수)
    cell.font = Font(bold=True, color='FFFFFF')
    # horizontal - 수직으로 가운데 정렬
    cell.alignment = Alignment(horizontal='center')
    # 배경을 단색 지정
    # : PatternFill은 그라이데션 색상을 지정
    # : start_color와 end_color를 동일하게 지정하는 경우 단색 지정이 가능
    cell.fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')

# 각 열의 너비를 조정
# 시트명.column_dimensions['열기호'].widht = 너비
sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 20
sheet.column_dimensions['C'].width = 20

# 행의 높이 조정
# 시트명.row_dimensions[행번호].height = 높이
sheet.row_dimensions[1].height = 20

# 동일한(존재하는) 파일을 업데이트
# : save('동일한파일명.확장자')
wb.save('example_modified.xlsx')

# 엑셀 파일 닫기
wb.close()